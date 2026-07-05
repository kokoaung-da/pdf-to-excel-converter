import os
import threading
import pdfplumber
import pandas as pd
import customtkinter as ctk
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# GUI Theme Settings
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class ProductionPDFExtractor(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Professional PDF Table Extractor")
        self.geometry("620x290")  # Folder Row ဖြုတ်လိုက်သဖြင့် Window အမြင့်ကို လျှော့ချထားပါသည်
        self.resizable(False, False)

        # Variables
        self.pdf_path = ctk.StringVar()

        # UI Layout
        self.title_label = ctk.CTkLabel(self, text="PDF to Excel Converter Pro", font=ctk.CTkFont(size=22, weight="bold"))
        self.title_label.pack(pady=(25, 20))

        # Main Grid Container for PDF Input Only
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="x", padx=30, pady=5)
        self.main_frame.columnconfigure(0, weight=1)  

        # Row 0: PDF Selection Only
        self.pdf_entry = ctk.CTkEntry(self.main_frame, textvariable=self.pdf_path, placeholder_text="Select PDF File...", height=35)
        self.pdf_entry.grid(row=0, column=0, padx=(0, 10), pady=10, sticky="ew")
        
        self.pdf_btn = ctk.CTkButton(self.main_frame, text="Browse PDF", command=self.browse_pdf, width=110, height=35, font=ctk.CTkFont(weight="bold"))
        self.pdf_btn.grid(row=0, column=1, pady=10, sticky="e")

        # Progress Status & Bar
        self.status_label = ctk.CTkLabel(self, text="Ready", font=ctk.CTkFont(size=12))
        self.status_label.pack(pady=(10, 5))
        
        self.progress_bar = ctk.CTkProgressBar(self, width=560, height=8)
        self.progress_bar.set(0)
        self.progress_bar.pack(pady=5)

        # Action Button
        self.convert_btn = ctk.CTkButton(self, text="Start Extraction", command=self.start_threading, fg_color="#1f85de", hover_color="#155e9c", width=220, height=42, font=ctk.CTkFont(size=14, weight="bold"))
        self.convert_btn.pack(pady=20)

    def browse_pdf(self):
        file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if file_path:
            self.pdf_path.set(file_path)

    def start_threading(self):
        pdf_file = self.pdf_path.get()

        if not pdf_file:
            messagebox.showwarning("Warning", "Please select a PDF file first!")
            return

        self.convert_btn.configure(state="disabled", text="Processing...")
        self.pdf_btn.configure(state="disabled")

        # Output Folder ကို PDF တည်ရှိရာ Folder အဖြစ် Auto သတ်မှတ်ခြင်း
        out_folder = os.path.dirname(pdf_file)

        threading.Thread(target=self.extract_pdf_process, args=(pdf_file, out_folder), daemon=True).start()

    def extract_pdf_process(self, pdf_file, out_folder):
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        pdf_base_name = os.path.splitext(os.path.basename(pdf_file))[0]
        file_name = f"{pdf_base_name}_{current_time}.xlsx"
        output_excel = os.path.join(out_folder, file_name)

        all_data_frames = []
        global_header = None
        total_tables_found = 0

        table_settings = {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "intersection_tolerance": 5
        }

        try:
            with pdfplumber.open(pdf_file) as pdf:
                total_pages = len(pdf.pages)
                
                for idx, page in enumerate(pdf.pages, start=1):
                    progress_val = idx / total_pages
                    self.progress_bar.set(progress_val)
                    self.status_label.configure(text=f"Processing page {idx} of {total_pages}...")
                    self.update_idletasks()

                    tables = page.extract_tables(table_settings)
                    
                    if not tables:
                        fallback_settings = {"vertical_strategy": "text", "horizontal_strategy": "text"}
                        tables = page.extract_tables(fallback_settings)

                    for table in tables:
                        if not table or all(len(row) == 0 for row in table):
                            continue
                        
                        if global_header is None:
                            global_header = [cell if cell else f"Col_{i}" for i, cell in enumerate(table[0])]
                            rows = table[1:]
                        else:
                            if table[0] == global_header or any(curr in global_header for curr in table[0] if curr):
                                rows = table[1:]
                            else:
                                rows = table

                        if not rows:
                            continue

                        df = pd.DataFrame(rows, columns=global_header if len(rows[0]) == len(global_header) else None)
                        df.dropna(how="all", inplace=True)
                        df.fillna("", inplace=True)
                        
                        all_data_frames.append(df)
                        total_tables_found += 1

            if all_data_frames:
                self.status_label.configure(text="Formatting Excel file...")
                self.update_idletasks()

                combined_df = pd.concat(all_data_frames, ignore_index=True)
                combined_df.to_excel(output_excel, index=False)

                # Excel Formatting
                wb = load_workbook(output_excel)
                ws = wb.active
                
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                wb.save(output_excel)

                success_msg = (
                    f"🎉 Extraction Completed!\n\n"
                    f"📄 PDF Name: {os.path.basename(pdf_file)}\n"
                    f"📖 Total Pages: {total_pages}\n"
                    f"📊 Tables Extracted: {total_tables_found}\n\n"
                    f"💾 Saved in Same Folder As:\n{file_name}"
                )
                messagebox.showinfo("Success", success_msg)
            else:
                messagebox.showwarning("No Tables", "⚠️ No structural tables could be extracted from this PDF.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during extraction:\n{str(e)}")
        
        finally:
            self.convert_btn.configure(state="normal", text="Start Extraction")
            self.pdf_btn.configure(state="normal")
            self.progress_bar.set(0)
            self.status_label.configure(text="Ready")

if __name__ == "__main__":
    app = ProductionPDFExtractor()
    app.mainloop()